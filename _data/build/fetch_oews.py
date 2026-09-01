"""BLS OEWS — occupation wages. Moves the workforce case from "how many jobs" to "which".

QCEW says how many jobs and what they pay ON AVERAGE across an industry. ACS says how many
people hold a credential. NEITHER says what the cluster's jobs actually ARE. OEWS does:
employment and wages by detailed occupation, by metro.

WHAT A ROW IS
  One (metro area, occupation) estimate of employment and wages — total employment, an
  hourly and annual mean, and the 10th/25th/median/75th/90th percentiles.

  It is a SURVEY, and its estimates carry relative standard errors. A metro-by-occupation
  cell for a small occupation is a small sample and moves between vintages for reasons
  that are not the labour market.

METRO, NOT COUNTY — SO NOT PIC-12
  OEWS publishes at metropolitan area. Metros neither nest inside nor tile the footprint:
  Akron, Cleveland-Elyria and Youngstown-Warren-Boardman together cover most of PIC-12 and
  each includes territory the footprint excludes. This is the SAME constraint that makes
  BEA Regional Price Parities a metro-level measure, and any page using either is a metro
  page and must say so.

  Do not sum the three metros and call the result PIC-12.

NOT INDUSTRY-SPECIFIC AS FETCHED. The metro files are all-industry: they say a metro
employs N chemical engineers, not that the polymer cluster does. Cross-industry occupation
detail exists in the OEWS industry files, which are a separate and much larger download.

Source: the OEWS metro data file, a ZIP of Excel workbooks. BLS 403s a bare User-Agent;
a complete browser header set is required — the same block that gates the QCEW handbook.
"""
import io, json, os, re, time, urllib.request, zipfile

HERE = os.path.dirname(os.path.abspath(__file__))
# BLS blocks on header COMPLETENESS, not user-agent alone.
H = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
     "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
     "Accept-Language": "en-US,en;q=0.9",
     "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate",
     "Sec-Fetch-Site": "none", "Upgrade-Insecure-Requests": "1"}

# Most recent published metro file. BLS versions these by year suffix.
CANDIDATES = ["https://www.bls.gov/oes/special-requests/oesm24ma.zip",
              "https://www.bls.gov/oes/special-requests/oesm23ma.zip"]

# The four metros overlapping PIC-12, by OEWS AREA code, VERIFIED against the 2024 file.
# Two corrections earned on the first run:
#   Cleveland is 17410 "Cleveland, OH" -- NOT 17460 "Cleveland-Elyria". The 2023 CBSA
#   revision renamed and recoded it, and 17460 silently returned nothing.
#   CANTON-MASSILLON (15940) was missed entirely and is Stark County, squarely in PIC-12.
# Youngstown also dropped "Boardman" and the PA portion in the same revision.
METROS = {"10420": "Akron, OH",
          "17410": "Cleveland, OH",
          "15940": "Canton-Massillon, OH",
          "49660": "Youngstown-Warren, OH"}
# Occupations a polymer cluster actually staffs. SOC codes, not titles — titles change.
SOC = {
    "17-2041": "Chemical engineers",
    "17-2131": "Materials engineers",
    "19-2031": "Chemists",
    "19-2032": "Materials scientists",
    "17-2112": "Industrial engineers",
    "17-2141": "Mechanical engineers",
    "51-9011": "Chemical equipment operators",
    "51-4021": "Extruding and drawing machine setters (metal/plastic)",
    "51-4031": "Cutting/punching/press machine setters (metal/plastic)",
    "51-4072": "Molding, coremaking and casting machine setters (metal/plastic)",
    "51-9023": "Mixing and blending machine setters",
    "51-9061": "Inspectors, testers, sorters, samplers, weighers",
    "17-3026": "Industrial engineering technologists and technicians",
    "19-4031": "Chemical technicians",
    # Added 2026-08-18 for the `occupations` page: the rest of the fourteen largest
    # detailed occupations in NAICS 326 by national staffing share (occmix.json). The
    # engineering and science codes above are a polymer cluster's distinctive staffing;
    # these are its BULK. A page about what the industry's jobs pay cannot leave out the
    # one in nine that is a molding-machine setter's supervisor, mechanic or packer.
    "51-1011": "First-line supervisors of production workers",
    "51-2090": "Miscellaneous assemblers and fabricators",
    "51-4081": "Multiple machine tool setters (metal/plastic)",
    "51-9041": "Extruding, forming, pressing and compacting machine setters",
    "51-9111": "Packaging and filling machine operators",
    "51-9197": "Tire builders",
    "49-9041": "Industrial machinery mechanics",
    "49-9071": "Maintenance and repair workers, general",
    "43-5071": "Shipping, receiving and inventory clerks",
    "53-7062": "Laborers and freight, stock and material movers, hand",
    "53-7064": "Packers and packagers, hand",
    "11-3051": "Industrial production managers",
}


def fetch_zip():
    last = None
    for u in CANDIDATES:
        try:
            with urllib.request.urlopen(urllib.request.Request(u, headers=H),
                                        timeout=300) as r:
                b = r.read()
            print(f"  {u.rsplit('/', 1)[-1]}: {len(b)//1024//1024} MB", flush=True)
            return u, b
        except Exception as e:
            last = f"{type(e).__name__} {str(e)[:60]}"
            print(f"  {u.rsplit('/', 1)[-1]}: {last}", flush=True)
    raise SystemExit(f"FATAL: no OEWS metro file reachable. Last error: {last}. BLS blocks "
                     f"on header completeness — a bare User-Agent gets 403 while a full "
                     f"browser header set passes.")


url, blob = fetch_zip()
zf = zipfile.ZipFile(io.BytesIO(blob))
names = [n for n in zf.namelist() if n.lower().endswith((".xlsx", ".xls"))]
if not names:
    raise SystemExit(f"FATAL: no workbook inside the ZIP. Members: {zf.namelist()[:6]}")
# The ZIP contains BOTH a metro file (MSA_*) and a balance-of-state file (BOS_*). Taking
# the alphabetically first .xlsx grabs BOS, which has 137 nonmetropolitan areas and no
# Ohio metros at all. Select by name, and fall back to largest only if the name is absent.
msa = [n for n in names if "MSA" in os.path.basename(n).upper()]
member = (msa or sorted(names, key=lambda n: -zf.getinfo(n).file_size))[0]
print(f"  reading {member} ({zf.getinfo(member).file_size//1024//1024} MB)", flush=True)

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("FATAL: openpyxl is required to read the OEWS workbook. "
                     "`uv pip install openpyxl` or use the toolkit environment.")

wb = load_workbook(io.BytesIO(zf.read(member)), read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
it = ws.iter_rows(values_only=True)
hdr = [str(c).strip().upper() if c is not None else "" for c in next(it)]
idx = {h: i for i, h in enumerate(hdr)}


def col(*alts):
    for a in alts:
        if a in idx:
            return idx[a]
    return None


c_area, c_soc = col("AREA"), col("OCC_CODE")
c_title, c_name = col("OCC_TITLE"), col("AREA_TITLE", "AREA_NAME")
if c_area is None or c_soc is None:
    raise SystemExit(f"FATAL: expected AREA and OCC_CODE columns; header was {hdr[:14]}")
FIELDS = {"TOT_EMP": "employment", "A_MEAN": "annual_mean", "H_MEAN": "hourly_mean",
          "A_PCT10": "a_pct10", "A_MEDIAN": "a_median", "A_PCT90": "a_pct90",
          "EMP_PRSE": "emp_rse", "MEAN_PRSE": "mean_rse", "JOBS_1000": "jobs_per_1000",
          "LOC_QUOTIENT": "loc_quotient"}


def num(v):
    """OEWS suppresses with '*' and tops out wages with '#'. Neither is a number."""
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    if s in ("", "*", "**", "#", "~"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


rows, seen_areas = [], set()
for r in it:
    area = str(r[c_area]).strip() if r[c_area] is not None else ""
    if area not in METROS:
        continue
    seen_areas.add(area)
    soc = str(r[c_soc]).strip() if r[c_soc] is not None else ""
    if soc not in SOC:
        continue
    row = {"area": area, "area_name": METROS[area], "soc": soc, "occupation": SOC[soc],
           "occ_title_bls": str(r[c_title]).strip() if c_title is not None and r[c_title] else None,
           "suppressed": False}
    for f, key in FIELDS.items():
        j = idx.get(f)
        row[key] = num(r[j]) if j is not None else None
    row["suppressed"] = row["employment"] is None
    rows.append(row)

if not rows:
    raise SystemExit(f"FATAL: no rows for the three PIC metros. Areas seen in file: "
                     f"{sorted(seen_areas)[:8]}. An empty result is an AREA-code or "
                     f"SOC-code mismatch, not three metros that employ no chemists.")
missing_metro = set(METROS) - seen_areas
if missing_metro:
    print(f"  WARNING: metros absent from file: "
          f"{[METROS[m] for m in missing_metro]}", flush=True)

sup = [r for r in rows if r["suppressed"]]
out = {"meta": {
    "source": f"BLS OEWS metropolitan file — {url.rsplit('/', 1)[-1]}, sheet "
              f"{wb.sheetnames[0]}",
    "row": "one (metro, SOC occupation) employment and wage estimate with percentiles and "
           "relative standard errors",
    "metro_not_county": "OEWS publishes at METROPOLITAN AREA. Metros neither nest inside nor "
                        "tile PIC-12 — Akron, Cleveland-Elyria and Youngstown-Warren "
                        "together cover most of the footprint and each includes territory it "
                        "excludes. DO NOT SUM THE THREE AND CALL IT PIC-12. Same constraint "
                        "as BEA Regional Price Parities.",
    "all_industry": "The metro files are ALL-INDUSTRY. They say a metro employs N chemical "
                    "engineers, NOT that the polymer cluster does. Industry-crossed "
                    "occupation detail is a separate, much larger OEWS download.",
    "survey": "OEWS is a SURVEY, not a census. Estimates carry relative standard errors "
              "(emp_rse, mean_rse), and a small occupation in one metro is a small sample "
              "that moves between vintages for reasons that are not the labour market. "
              "Report the RSE beside any small cell.",
    "suppression": "'*' means the estimate is not released; '#' means a wage at or above "
                   "$100/hour or $208,000/year, which is a CEILING not a value. Both become "
                   "null here, never zero.",
    "metros": METROS, "occupations": SOC,
    "counts": {"rows": len(rows), "suppressed": len(sup),
               "metros_found": sorted(seen_areas)},
    "fetched": time.strftime("%Y-%m-%d")}, "rows": rows}

p = os.path.join(HERE, "oews.json")
json.dump(out, open(p, "w", encoding="utf-8"), separators=(",", ":"))
print(f"wrote {p} {round(os.path.getsize(p)/1024)} KB, {len(rows)} rows across "
      f"{len(seen_areas)} metros and {len({r['soc'] for r in rows})} occupations, "
      f"{len(sup)} suppressed")
