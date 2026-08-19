"""BLS OEWS — the NATIONAL benchmark for the metro occupation wages in oews.json.

WHY A SECOND FILE. `fetch_oews.py` pulls the metropolitan workbook, which says what a
metro pays a chemical engineer. It says nothing about what the country pays one, and a
metro median without a national one beside it is a number with no bar. This pulls the
national workbook for the SAME occupations, so the two files always describe one set.

WHAT A ROW IS
  One (national, SOC occupation) estimate: total employment, hourly and annual mean, and
  the 10th/median/90th annual percentiles, with relative standard errors. All industries.
  A SURVEY, with the same '*' (not released) and '#' (wage at or above the ceiling)
  conventions as the metro file; both become null here, never zero.

THE OCCUPATION SET IS READ FROM oews.json, not typed here. The metro pull is the source of
truth for which occupations the site follows; if a code is added there it appears here on
the next run and cannot silently diverge. Run order: fetch_oews.py, then this.

Source: the OEWS national data file, a ZIP of one Excel workbook (~2 MB). BLS 403s a bare
User-Agent; the same complete browser header set as fetch_oews.py is required.
"""
import io, json, os, time, urllib.request, zipfile

HERE = os.path.dirname(os.path.abspath(__file__))
H = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
     "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
     "Accept-Language": "en-US,en;q=0.9",
     "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate",
     "Sec-Fetch-Site": "none", "Upgrade-Insecure-Requests": "1"}
CANDIDATES = ["https://www.bls.gov/oes/special-requests/oesm24nat.zip",
              "https://www.bls.gov/oes/special-requests/oesm23nat.zip"]

metro_path = os.path.join(HERE, "oews.json")
if not os.path.exists(metro_path):
    raise SystemExit("FATAL: oews.json is missing — run fetch_oews.py first; the occupation "
                     "set is read from it so the two files can never disagree.")
metro = json.load(open(metro_path, encoding="utf-8"))
SOC = metro["meta"]["occupations"]
metro_vintage = metro["meta"]["source"]


def fetch_zip():
    last = None
    for u in CANDIDATES:
        try:
            with urllib.request.urlopen(urllib.request.Request(u, headers=H),
                                        timeout=300) as r:
                b = r.read()
            print(f"  {u.rsplit('/', 1)[-1]}: {len(b)//1024} KB", flush=True)
            return u, b
        except Exception as e:
            last = f"{type(e).__name__} {str(e)[:60]}"
            print(f"  {u.rsplit('/', 1)[-1]}: {last}", flush=True)
    raise SystemExit(f"FATAL: no OEWS national file reachable. Last error: {last}.")


url, blob = fetch_zip()
if url.rsplit("/", 1)[-1][4:6] not in metro_vintage:
    raise SystemExit(f"FATAL: national file {url.rsplit('/', 1)[-1]} is a different vintage "
                     f"from the metro pull ({metro_vintage}). Re-run fetch_oews.py.")
zf = zipfile.ZipFile(io.BytesIO(blob))
names = [n for n in zf.namelist() if n.lower().endswith((".xlsx", ".xls"))]
if not names:
    raise SystemExit(f"FATAL: no workbook inside the ZIP. Members: {zf.namelist()[:6]}")
member = sorted(names, key=lambda n: -zf.getinfo(n).file_size)[0]
print(f"  reading {member} ({zf.getinfo(member).file_size//1024} KB)", flush=True)

from openpyxl import load_workbook
wb = load_workbook(io.BytesIO(zf.read(member)), read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
it = ws.iter_rows(values_only=True)
hdr = [str(c).strip().upper() if c is not None else "" for c in next(it)]
idx = {h: i for i, h in enumerate(hdr)}
c_soc, c_title, c_group = idx.get("OCC_CODE"), idx.get("OCC_TITLE"), idx.get("O_GROUP")
if c_soc is None:
    raise SystemExit(f"FATAL: expected OCC_CODE; header was {hdr[:14]}")
FIELDS = {"TOT_EMP": "employment", "A_MEAN": "annual_mean", "H_MEAN": "hourly_mean",
          "A_PCT10": "a_pct10", "A_MEDIAN": "a_median", "A_PCT90": "a_pct90",
          "EMP_PRSE": "emp_rse", "MEAN_PRSE": "mean_rse"}


def num(v):
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    if s in ("", "*", "**", "#", "~"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


rows = []
for r in it:
    soc = str(r[c_soc]).strip() if r[c_soc] is not None else ""
    if soc not in SOC:
        continue
    row = {"soc": soc, "occupation": SOC[soc],
           "occ_title_bls": str(r[c_title]).strip() if c_title is not None and r[c_title] else None,
           "o_group": str(r[c_group]).strip() if c_group is not None and r[c_group] else None}
    for f, key in FIELDS.items():
        j = idx.get(f)
        row[key] = num(r[j]) if j is not None else None
    row["suppressed"] = row["a_median"] is None
    rows.append(row)

# 51-2090 is listed twice in the 2024 national workbook — once as a broad group and once
# as the detailed line the 2018 SOC collapsed it to — with identical values. One row per
# code, detailed preferred; two identical rows would double the occupation in any join.
by_soc = {}
for r in rows:
    if r["soc"] not in by_soc or (r["o_group"] == "detailed" and by_soc[r["soc"]]["o_group"] != "detailed"):
        by_soc[r["soc"]] = r
rows = list(by_soc.values())
missing = sorted(set(SOC) - {r["soc"] for r in rows})
if len(rows) < len(SOC) - 2:
    raise SystemExit(f"FATAL: only {len(rows)} of {len(SOC)} occupations found nationally; "
                     f"missing {missing}. A national file that lacks an occupation the "
                     f"metro file has is a code mismatch, not a fact.")
out = {"meta": {
    "source": f"BLS OEWS national file — {url.rsplit('/', 1)[-1]}, sheet {wb.sheetnames[0]}",
    "row": "one (national, SOC occupation) all-industry employment and wage estimate with "
           "percentiles and relative standard errors",
    "all_industry": "ALL INDUSTRIES. The national median for a chemist is the median across "
                    "every industry that employs chemists, not the plastics industry's.",
    "suppression": "'*' means not released; '#' means a wage at or above $115/hour or "
                   "$239,200/year, a CEILING not a value. Both become null, never zero.",
    "occupations": SOC, "counts": {"rows": len(rows), "missing": missing},
    "fetched": time.strftime("%Y-%m-%d")}, "rows": rows}
p = os.path.join(HERE, "oews_national.json")
json.dump(out, open(p, "w", encoding="utf-8"), separators=(",", ":"))
print(f"wrote {p} {round(os.path.getsize(p)/1024)} KB, {len(rows)} occupations"
      + (f", missing {missing}" if missing else ""))
